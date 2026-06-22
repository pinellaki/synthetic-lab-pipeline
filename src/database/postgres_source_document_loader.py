"""Load raw source-document metadata into PostgreSQL.

This module scans the local data/raw directory and records metadata about
the files that enter the pipeline.

It tracks CSV files, Excel workbooks, API JSON pages, PDF reports, and text
reports. The goal is to support data-lineage and governance dashboard views:
what entered the pipeline, what type of source it was, and how many records,
pages, rows, or lines were detected.

The loader is safe to rerun. It uses an upsert on source_path, so existing
source-document rows are updated instead of duplicated.
"""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader

from src.database.postgres_connection import create_postgres_connection


PROJECT_ROOT = Path(__file__).resolve().parents[2]
RAW_DATA_ROOT = PROJECT_ROOT / "data" / "raw"


@dataclass(frozen=True)
class SourceDocumentRow:
    """Metadata about one raw source file."""

    source_path: str
    file_name: str
    source_type: str
    file_extension: str
    file_size_bytes: int
    records_detected: int | None
    records_loaded: int | None
    records_rejected: int | None
    ingestion_status: str
    notes: str


def classify_source_file(path: Path) -> str:
    """Classify a raw source file into a dashboard-friendly source type."""
    extension = path.suffix.lower()
    path_parts = {part.lower() for part in path.parts}

    if extension == ".csv":
        return "CSV"

    if extension in {".xlsx", ".xls"}:
        return "EXCEL"

    if extension == ".json":
        if "api_pages" in path_parts or "api" in path.stem.lower():
            return "API_JSON"
        return "API_JSON"

    if extension == ".pdf":
        return "PDF_REPORT"

    if extension == ".txt":
        return "TEXT_REPORT"

    return "OTHER"


def count_csv_records(path: Path) -> int:
    """Count data rows in a CSV file, excluding the header row."""
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        return sum(1 for _ in reader)


def count_excel_records(path: Path) -> int:
    """Count data rows across all sheets in an Excel workbook."""
    workbook = load_workbook(path, read_only=True, data_only=True)

    try:
        total_rows = 0

        for worksheet in workbook.worksheets:
            if worksheet.max_row and worksheet.max_row > 1:
                total_rows += worksheet.max_row - 1

        return total_rows

    finally:
        workbook.close()


def count_json_records(path: Path) -> int:
    """Count records in a JSON file.

    If the root is a list, each item is treated as one record.
    If the root is a dictionary with a common list key, that list is counted.
    Otherwise, the file is counted as one detected payload.
    """
    with path.open("r", encoding="utf-8") as file:
        payload: Any = json.load(file)

    if isinstance(payload, list):
        return len(payload)

    if isinstance(payload, dict):
        for key in ("data", "results", "items", "records", "shipments"):
            value = payload.get(key)

            if isinstance(value, list):
                return len(value)

        return 1

    return 1


def count_pdf_pages(path: Path) -> int:
    """Count pages in a PDF report."""
    reader = PdfReader(str(path))
    return len(reader.pages)


def count_text_lines(path: Path) -> int:
    """Count non-empty lines in a text report."""
    with path.open("r", encoding="utf-8", errors="ignore") as file:
        return sum(1 for line in file if line.strip())


def detect_record_count(path: Path, source_type: str) -> tuple[int | None, str | None]:
    """Return detected count and optional error note for one file."""
    try:
        if source_type == "CSV":
            return count_csv_records(path), None

        if source_type == "EXCEL":
            return count_excel_records(path), None

        if source_type == "API_JSON":
            return count_json_records(path), None

        if source_type == "PDF_REPORT":
            return count_pdf_pages(path), None

        if source_type == "TEXT_REPORT":
            return count_text_lines(path), None

        return None, "File detected, but no record-count strategy exists."

    except Exception as exc:
        return None, f"File detected, but count failed: {exc}"


def build_notes(source_type: str, count_error: str | None) -> str:
    """Build a clear governance note for a source file."""
    if count_error:
        return count_error

    if source_type in {"CSV", "EXCEL", "API_JSON"}:
        return (
            "Structured source detected. This source type can be used by the "
            "pipeline for loading trusted PostgreSQL tables."
        )

    if source_type in {"PDF_REPORT", "TEXT_REPORT"}:
        return (
            "Raw report detected. The file is tracked as source metadata for "
            "governance and lineage; report content is not loaded as a trusted "
            "domain record in the current one-time loader."
        )

    return "Raw source detected and tracked as metadata."


def choose_ingestion_status(source_type: str, count_error: str | None) -> str:
    """Choose a simple ingestion status for the source-document table."""
    if count_error:
        return "error"

    if source_type in {"CSV", "EXCEL", "API_JSON"}:
        return "processed"

    if source_type in {"PDF_REPORT", "TEXT_REPORT", "OTHER"}:
        return "detected_only"

    return "detected_only"


def build_source_document_rows(
    raw_data_root: Path = RAW_DATA_ROOT,
) -> list[SourceDocumentRow]:
    """Scan data/raw and build source-document rows."""
    if not raw_data_root.exists():
        return []

    rows: list[SourceDocumentRow] = []

    for path in sorted(raw_data_root.rglob("*")):
        if not path.is_file():
            continue

        if path.name == ".gitkeep":
            continue

        source_type = classify_source_file(path)
        records_detected, count_error = detect_record_count(path, source_type)
        relative_path = path.relative_to(PROJECT_ROOT).as_posix()

        rows.append(
            SourceDocumentRow(
                source_path=relative_path,
                file_name=path.name,
                source_type=source_type,
                file_extension=path.suffix.lower(),
                file_size_bytes=path.stat().st_size,
                records_detected=records_detected,
                records_loaded=None,
                records_rejected=None,
                ingestion_status=choose_ingestion_status(
                    source_type=source_type,
                    count_error=count_error,
                ),
                notes=build_notes(
                    source_type=source_type,
                    count_error=count_error,
                ),
            )
        )

    return rows


def upsert_source_documents(rows: list[SourceDocumentRow]) -> int:
    """Insert or update source-document metadata in PostgreSQL."""
    if not rows:
        return 0

    with create_postgres_connection() as connection:
        with connection.cursor() as cursor:
            for row in rows:
                cursor.execute(
                    """
                    INSERT INTO source_document (
                        source_path,
                        file_name,
                        source_type,
                        file_extension,
                        file_size_bytes,
                        records_detected,
                        records_loaded,
                        records_rejected,
                        ingestion_status,
                        notes
                    )
                    VALUES (
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s
                    )
                    ON CONFLICT (source_path)
                    DO UPDATE SET
                        file_name = EXCLUDED.file_name,
                        source_type = EXCLUDED.source_type,
                        file_extension = EXCLUDED.file_extension,
                        file_size_bytes = EXCLUDED.file_size_bytes,
                        records_detected = EXCLUDED.records_detected,
                        records_loaded = EXCLUDED.records_loaded,
                        records_rejected = EXCLUDED.records_rejected,
                        ingestion_status = EXCLUDED.ingestion_status,
                        notes = EXCLUDED.notes,
                        updated_at = CURRENT_TIMESTAMP
                    """,
                    (
                        row.source_path,
                        row.file_name,
                        row.source_type,
                        row.file_extension,
                        row.file_size_bytes,
                        row.records_detected,
                        row.records_loaded,
                        row.records_rejected,
                        row.ingestion_status,
                        row.notes,
                    ),
                )

        connection.commit()

    return len(rows)


def load_source_documents() -> int:
    """Scan raw files and upsert their metadata into PostgreSQL."""
    rows = build_source_document_rows()
    return upsert_source_documents(rows)


if __name__ == "__main__":
    loaded_count = load_source_documents()
    print(f"Source-document metadata rows loaded: {loaded_count}")